# data_loader_streamlit.py
import os
import glob
import pandas as pd
import numpy as np
import calendar
import streamlit as st
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from excel_formatter import *

class DataLoader:
    def __init__(self, folder_path: str):
        self.folder_path = folder_path
        self.file_name = "In-Out-Stock-History.xlsx"
        self.init_dataframes()
        self.init_agg_dicts()

    # ---------- INITIALIZE ----------
    def init_dataframes(self):
        self.df_combined_csv = None
        self.df_receive_ship_qty = None
        self.df_stock_on_receive_ship_date = None
        self.df_stock_not_zero = None
        self.df_stock_on_receive_ship_and_stock_not_zero = None
        self.df_daily_stock = None
        self.df_receive_ship_stock_daily = None
        self.df_daily_transaction = None
        self.df_monthly_transaction = None
        self.df_weekly_transaction = None
        self.df_yearly_transaction = None
        self.df_stock_aging = None    
        self.df_storage_day = None

    def init_agg_dicts(self):
        self.agg_dict1 = {
            'Quantity[Unit1]': 'sum',
            'Owner Name': 'first',
            'Item Name': 'first',
            'UOM1': 'first',
            'Delivery Destination Code': 'first',
            'Delivery Destination Name': 'first',      
            'Rcv So Flag': 'first'
        }
        self.agg_dict2 = {
            'Quantity[Unit1]': 'sum',
            'Owner Name': 'first',
            'Item Name': 'first',
            'UOM1': 'first',
            'Delivery Destination Code': 'first',
            'Delivery Destination Name': 'first',                  
        }
        self.agg_dict3 = {
            'Quantity[Unit1]': 'first',
            'Owner Name': 'first',
            'Item Name': 'first',
            'UOM1': 'first',
            'Delivery Destination Code': 'first',
            'Delivery Destination Name': 'first',
            'Rcv So Flag': 'first'            
        }    

    # ---------- MAIN PIPELINE ----------
    def run(self):
        progress = st.progress(0)
        status = st.empty()
        step = 0
        steps_total = 12

        def update_status(msg):
            nonlocal step
            step += 1
            status.text(f"[{step}/{steps_total}] {msg}")
            progress.progress(int(step/steps_total*100))

        update_status("Loading CSV files...")
        self.load_csv_file()
        update_status("Aggregating receive/ship qty...")
        self.receive_ship_qty()
        update_status("Filtering non-zero stock...")
        self.stock_not_zero()
        update_status("Computing stock on receive/ship date...")
        self.stock_on_receive_ship_date()
        update_status("Combining stock info...")
        self.stock_on_receive_ship_and_stock_not_zero()
        update_status("Resampling daily stock...")
        self.resample_daily_stock()
        update_status("Creating receive/ship stock daily...")
        self.receive_ship_stock_daily()
        update_status("Building daily transaction...")
        self.daily_transaction()
        update_status("Building weekly transaction...")
        self.weekly_transaction()
        update_status("Building monthly transaction...")
        self.monthly_transaction()
        update_status("Building yearly transaction...")
        self.yearly_transaction()
        update_status("Calculating stock aging...")
        self.stock_aging()
        update_status("Calculating storage day...")
        self.storage_day()

        update_status("Saving to Excel...")
        self.save_to_excel()
        progress.progress(100)
        status.text("✅ Completed")

    # ---------- METHODS (เดิมทั้งหมด) ----------
    def load_csv_file(self):
        csv_files = glob.glob(os.path.join(self.folder_path, "**", "*.csv"), recursive=True)
        df_list = []
        for file in csv_files:
            usecols = ["Operation Date", "Rcv So Flag", "Owner Code", "Owner Name", "Item Code", "Item Name",
                       "Quantity[Unit1]", "UOM1", "Inventory Qty", "Delivery Destination Code", "Delivery Destination Name"]
            df = pd.read_csv(file, usecols=usecols, dtype={'Item Code': 'str', 'Item Name': 'str'})
            df_list.append(df)

        self.df_combined_csv = pd.concat(df_list, ignore_index=True)
        self.df_combined_csv['Rcv So Flag'] = self.df_combined_csv['Rcv So Flag'].replace({
            "Rcv(increase)": "In",
            "So(decrese)": "Out"
        })
        self.df_combined_csv['Operation Date'] = pd.to_datetime(
            self.df_combined_csv['Operation Date'], dayfirst=True, errors='coerce'
        )
        self.df_combined_csv = self.df_combined_csv.dropna(subset=['Operation Date', 'Quantity[Unit1]'])

    def receive_ship_qty(self):
        self.df_receive_ship_qty = (
            self.df_combined_csv
            .groupby(['Owner Code', 'Item Code', 'Operation Date', 'Rcv So Flag'], as_index=False)
            .agg(self.agg_dict2)
        )

    def stock_on_receive_ship_date(self):
        self.df_stock_on_receive_ship_date = (
            self.df_combined_csv
            .groupby(['Owner Code', 'Item Code', 'Operation Date'], as_index=False)
            .agg(self.agg_dict1)
        )
        self.df_stock_on_receive_ship_date['Rcv So Flag'] = "Stock"
        self.df_stock_on_receive_ship_date = self.df_stock_on_receive_ship_date.sort_values(
            ['Owner Code', 'Item Code']
        )
        self.df_stock_on_receive_ship_date['Quantity[Unit1]'] = (
            self.df_stock_on_receive_ship_date
            .groupby(['Owner Code', 'Item Code'])['Quantity[Unit1]']
            .cumsum()
        )

    def stock_not_zero(self):
        max_date = self.df_combined_csv['Operation Date'].max().replace(day=1) + pd.offsets.MonthEnd(0)
        self.df_stock_not_zero = (
            self.df_combined_csv
            .groupby(['Owner Code', 'Item Code'], as_index=False)
            .agg(self.agg_dict1)
        )
        self.df_stock_not_zero = self.df_stock_not_zero[self.df_stock_not_zero['Quantity[Unit1]']!=0]
        self.df_stock_not_zero['Rcv So Flag'] = 'Stock'
        self.df_stock_not_zero['Operation Date'] = max_date    

    def stock_on_receive_ship_and_stock_not_zero(self):
        self.df_stock_on_receive_ship_and_stock_not_zero = pd.concat([
            self.df_stock_on_receive_ship_date, self.df_stock_not_zero
        ])
        self.df_stock_on_receive_ship_and_stock_not_zero = (
            self.df_stock_on_receive_ship_and_stock_not_zero
            .groupby(['Owner Code', 'Item Code', 'Operation Date'],as_index=False)
            .agg(self.agg_dict3)
        )

    def resample_daily_stock(self):
        self.df_daily_stock = (
            self.df_stock_on_receive_ship_and_stock_not_zero
            .set_index('Operation Date')
            .groupby(['Owner Code', 'Item Code'], group_keys=False)
            .resample('D')
            .ffill()
            .reset_index()
        )

    def receive_ship_stock_daily(self):
        df_in = self.df_daily_stock.copy()
        df_in['Rcv So Flag'] = "In"
        df_in['Quantity[Unit1]'] = 0
        df_out = self.df_daily_stock.copy()
        df_out['Rcv So Flag'] = "Out"
        df_out['Quantity[Unit1]'] = 0
        self.df_receive_ship_stock_daily = pd.concat([
            self.df_receive_ship_qty, self.df_daily_stock, df_in, df_out
        ]).drop_duplicates().reset_index(drop=True)        
        self.df_receive_ship_stock_daily = (
            self.df_receive_ship_stock_daily
            .groupby(['Owner Code', 'Item Code', 'Operation Date', 'Rcv So Flag'], as_index=False)
            .agg(self.agg_dict2)
        )

    # ---------- TRANSACTION METHODS ----------
    def daily_transaction(self):
        self.df_daily_transaction = (
            self.df_receive_ship_stock_daily
            .groupby(['Owner Code', 'Item Code', 'Operation Date', 'Rcv So Flag'], as_index=False)
            .agg(self.agg_dict2)
        )

    def weekly_transaction(self):
        df = self.df_receive_ship_stock_daily.copy()
        df["Week"] = df["Operation Date"].dt.to_period("W").apply(lambda r: r.start_time)
        self.df_weekly_transaction = (
            df.groupby(['Owner Code', 'Item Code', 'Week', 'Rcv So Flag'], as_index=False)
              .agg(self.agg_dict2)
        )
        self.df_weekly_transaction.rename(columns={"Week": "Operation Date"}, inplace=True)

    def monthly_transaction(self):
        df = self.df_receive_ship_stock_daily.copy()
        df["Month"] = df["Operation Date"].dt.to_period("M").apply(lambda r: r.start_time)
        self.df_monthly_transaction = (
            df.groupby(['Owner Code', 'Item Code', 'Month', 'Rcv So Flag'], as_index=False)
              .agg(self.agg_dict2)
        )
        self.df_monthly_transaction.rename(columns={"Month": "Operation Date"}, inplace=True)

    def yearly_transaction(self):
        df = self.df_receive_ship_stock_daily.copy()
        df["Year"] = df["Operation Date"].dt.to_period("Y").apply(lambda r: r.start_time)
        self.df_yearly_transaction = (
            df.groupby(['Owner Code', 'Item Code', 'Year', 'Rcv So Flag'], as_index=False)
              .agg(self.agg_dict2)
        )
        self.df_yearly_transaction.rename(columns={"Year": "Operation Date"}, inplace=True)

    # ---------- STOCK AGING & STORAGE DAY ----------
    def stock_aging(self):
        # copy stock daily
        df = self.df_stock_daily.copy()
        df["Stock Age (Days)"] = (pd.Timestamp.today().normalize() - df["Operation Date"]).dt.days
        df.loc[df["Stock Age (Days)"] < 0, "Stock Age (Days)"] = 0
        self.df_stock_aging = df

    def storage_day(self):
        # Calculate storage days = Stock Qty / Movement
        df = self.df_stock_daily.copy()

        df["Movement"] = df["Rcv Qty"] + df["Ship Qty"]
        df["Storage Days"] = df.apply(
            lambda row: row["Stock Qty"] / row["Movement"] if row["Movement"] > 0 else None,
            axis=1
        )
        self.df_storage_day = df



    
    # ---------- SAVE TO EXCEL ----------
    def save_df_to_sheet_in_chunks(self, writer, df, sheet_name, chunk_size=1000):
        """Save DataFrame to Excel in chunks (no signals needed for Streamlit)."""
        wb = writer.book
        ws = wb.create_sheet(title=sheet_name)

        # Write header
        for row in dataframe_to_rows(df.iloc[:0], index=False, header=True):
            ws.append(row)

        total_rows = len(df)
        if total_rows == 0:
            return

        for i in range(0, total_rows, chunk_size):
            chunk = df.iloc[i:i + chunk_size]
            for row in dataframe_to_rows(chunk, index=False, header=False):
                ws.append(row)

        # Apply formatting
        format_sheet_column_width(ws)
        format_sheet_header_border(ws)
        format_sheet_header_horizontal_alignment(ws)
        format_sheet_header_font_bold(ws)
        format_sheet_detail_border(ws)
        format_sheet_detail_horizontal_alignment(ws)
        format_sheet_detail_number_format(ws)
        format_sheet_bottom_total_formula(ws)
        format_sheet_bottom_border(ws)
        format_sheet_add_auto_filter(ws)

    def save_to_excel(self):
        """Save all DataFrames into one Excel file."""
        try:
            with pd.ExcelWriter(self.file_name, engine='openpyxl') as writer:
                # Initialize workbook with dummy sheet
                pd.DataFrame().to_excel(writer, sheet_name='dummy', index=False)

                # List of all DataFrames to save
                sheets = [
                    ('Daily Transaction', self.df_daily_transaction),
                    ('Weekly Transaction', self.df_weekly_transaction),
                    ('Monthly Transaction', self.df_monthly_transaction),
                    ('Yearly Transaction', self.df_yearly_transaction),
                    ('Stock Aging', self.df_stock_aging),
                    ('Storage Day', self.df_storage_day),
                ]

                for sheet_name, df in sheets:
                    self.save_df_to_sheet_in_chunks(writer, df, sheet_name)

                # Remove dummy sheet
                if 'dummy' in writer.book.sheetnames:
                    writer.book.remove(writer.book['dummy'])

        except Exception as e:
            raise RuntimeError(f"Error saving Excel: {e}")




    # ---------- MAIN RUN ----------
    def run(self):
        """Run the full pipeline: load → aggregate → save → update session_state."""
        try:
            # Step 1: Load and preprocess raw data
            self.load_raw_data()

            # Step 2: Aggregations
            self.aggregate_daily()
            self.aggregate_weekly()
            self.aggregate_monthly()
            self.aggregate_yearly()

            # Step 3: Stock calculations
            self.calculate_stock_aging()
            self.calculate_storage_day()

            # Step 4: Save to Excel
            self.save_to_excel()

            # Step 5: Push into Streamlit session_state
            st.session_state["official_data"] = self.df_monthly_transaction
            st.session_state["selected_site"] = (
                self.df_monthly_transaction["Site"].dropna().unique()[0]
                if not self.df_monthly_transaction.empty else None
            )
            st.session_state["message"] = "✅ Data successfully loaded and processed."

        except Exception as e:
            st.error(f"❌ Error during run: {e}")




    # ---------- DOWNLOAD EXCEL ----------
    def render_download_button(self):
        """Render a Streamlit download button for the processed Excel file."""
        if self.output_path.exists():
            with open(self.output_path, "rb") as f:
                st.download_button(
                    label="📥 Download Processed Excel",
                    data=f,
                    file_name=self.output_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_processed_excel"
                )
        else:
            st.info("ℹ️ Processed Excel file not found. Please run the pipeline first.")

