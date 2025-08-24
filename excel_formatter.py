# excel_formatter.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO

class ExcelFormatter:
    def __init__(self, df: pd.DataFrame, file_name: str = "report.xlsx"):
        self.df = df
        self.file_name = file_name

    def save_to_excel(self, sheets: dict = None) -> BytesIO:
        """
        Export DataFrame(s) to Excel with basic formatting.
        sheets : dict of {sheet_name: dataframe}
        return : BytesIO (for Streamlit download button)
        """
        output = BytesIO()

        if sheets is None:
            sheets = {"Data": self.df}

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, data in sheets.items():
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)

        # --- formatting step ---
        output.seek(0)
        wb = load_workbook(output)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Header style
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")

            # Column auto width
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[col_letter].width = adjusted_width

        # Save formatted file
        output2 = BytesIO()
        wb.save(output2)
        output2.seek(0)
        return output2
